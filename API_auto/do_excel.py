# _*_ coding: utf_8 _*_
# @Time     : 2019/3/12
# @Author   : yuandian
# @Email    : 919839065@qq.com
# @file     : class_0312

import openpyxl


COOKIES = None

class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class do_excel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.work_book = openpyxl.load_workbook(file_name)
        self.sheet = self.work_book[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row
        cases = []
        for r in range(2, max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(row=r, column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.method = self.sheet.cell(row=r, column=4).value
            case.expected = self.sheet.cell(row=r, column=5).value
            case.actual = self.sheet.cell(row=r, column=6).value

            cases.append(case)

            self.work_book.close()

        return cases

    def write_data(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.work_book.save(filename=self.sheet_name)
        self.work_book.close()


if __name__ == '__main__':
    de = do_excel("python.xlsx", sheet_name="Sheet1")
    cases = de.get_cases()
    # global COOKIES
    # # rq = Case().request()
    # for case in cases:
    #     print(case.__dict__)
    #     resp = request(url=case.url, param=case.data, cookies = COOKIES)
    #     actual = resp.text
    #     print(actual)