# _*_ coding: utf_8 _*_
# @Time     : 2019/3/12
# @Author   : yuandian
# @Email    : 919839065@qq.com
# @file     : class_0312
from openpyxl import load_workbook
from API_auto.API_1 import file_name
from API_auto.API_1.HTTP_request import *


class DoExcel:
    def Get_date(self):
        wb = load_workbook(file_name.case_file)  # 打开excel
        CP = ConfigParsers()
        num = CP.get("sheet", "sheet_no")
        num = int(num)
        sheet = wb.worksheets[num]  # 定位到表单
        null_list_2 = []
        for i in range(2, sheet.max_row + 1):
            user = {
                "case_id": sheet.cell(i, 1).value,
                "title": sheet.cell(i, 2).value,
                "url": sheet.cell(i, 3).value,
                "data": sheet.cell(i, 4).value,
                "method": sheet.cell(i, 5).value,
                "expected": sheet.cell(i, 6).value,

            }
            null_list_2.append(user)
        return null_list_2

    def write_data(self, actuel, case_id):
        wb = load_workbook(file_name.case_file)
        CP = ConfigParsers()
        num = CP.get("sheet", "sheet_no")
        num = int(num)
        sheet = wb.worksheets[num]
        sheet.cell(case_id + 1, 7).value = actuel.json()["msg"]

        if sheet.cell(case_id + 1, 6).value == sheet.cell(case_id + 1, 7).value:
            sheet.cell(case_id + 1, 8).value = "PASS"
        else:
            sheet.cell(case_id + 1, 8).value = "FAIL"
        wb.save(file_name.case_file)  # 保存excel
        wb.close()  # 关闭excel
