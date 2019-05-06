# _*_ coding: utf_8 _*_
# @Time     : 2019/3/12
# @Author   : yuandian
# @Email    : 919839065@qq.com
# @file     : class_0312

from openpyxl import load_workbook
from week6.get_data import request

class Write_data:
    def write_data(self,filename,url,param):
        wb = load_workbook(filename)
        sheet = wb.worksheets[0]
        sheet.cell(2, 5).value =request(url,param).json()["msg"]
        print(sheet.cell(2,5).value)
        wb.save(filename)  # 保存excel
        wb.close()  # 关闭excel


if __name__ == '__main__':
    url = "http://47.107.168.87:8080/futureloan/mvc/api/member/login"
    param = {'mobilephone': '18688773467', 'pwd': '123456'}
    wd = Write_data()
    wd.write_data("python.xlsx",url,param)