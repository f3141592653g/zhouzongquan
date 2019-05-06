# _*_ coding: utf_8 _*_
# @Time     : 2019/3/12
# @Author   : yuandian
# @Email    : 919839065@qq.com
# @file     : class_0312
from openpyxl import load_workbook
import requests


def Get_date():
    wb = load_workbook("python.xlsx")  # 打开excel
    sheet = wb.worksheets[0]  # 定位到表单
    null_list_2 = []
    for i in range(2, sheet.max_row + 1):
        null_list_1 = []
        for j in range(1, 5):
            st = sheet.cell(i, j).value
            null_list_1.append(st)
        null_list_2.append(null_list_1)
    wb.close()  # 关闭excel
    return null_list_2


def request(url, param, method="GET"):
    if method.upper() == 'GET':
        result = requests.get(url, param)
    elif method.upper() == 'POST':
        result = requests.get(url, param)
    return result


