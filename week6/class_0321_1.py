# _*_ coding: utf_8 _*_
# @Time     : 2019/3/12
# @Author   : yuandian
# @Email    : 919839065@qq.com
# @file     : class_0312

from openpyxl import load_workbook

class work_book:
    def read_date(self,fileName):
        wb= load_workbook(fileName) #打开excel
        sheet = wb["Sheet1"] #定位到表单
        null_list_2 = []
        for i in range(1,sheet.max_row+1):
            null_list_1 = []
            for j in range(1,sheet.max_column+1):
                st = sheet.cell(i, j).value
                null_list_1.append(st)
            null_list_2.append(null_list_1)
            wb.close()  # 关闭excel
        return null_list_2


if __name__ == '__main__':
    wk = work_book().read_date("python.xlsx")
    print(wk)
