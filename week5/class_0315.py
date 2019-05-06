# _*_ coding: utf_8 _*_
# @Time     : 2019/3/15
# @Author   : yuandian
# @Email    : 919839065@qq.com
# @file     : class_0315

#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据

#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面,要求把读取到的数据返回 #2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值



#温馨提示：记得关闭和保存Excel

from openpyxl import load_workbook


class work_book:
    def read_date(self,fileName):
        wb= load_workbook(fileName) #打开excel
        sheet = wb["Sheet1"] #定位到表单
        null_list_2 = []
        for i in range(1,sheet.max_row+1):
            null_list_1 = []
            for j in range(1,sheet.max_column+1):
                st = sheet.cell(i, j).value
                null_list_1.append(st)
            null_list_2.append(null_list_1)
            wb.close()  # 关闭excel
        return null_list_2

    def write_date(self,fileName):
        wb=load_workbook(fileName)#打开excel
        sheet = wb["Sheet1"]
        sheet.cell(3,2).value='你很牛皮'
        print("第二题结果是：\n{}".format(sheet.cell(3,2).value))
        wb.save(fileName)#保存excel
        wb.close()  # 关闭excel


if __name__ == '__main__':
    wk=work_book()
    res = wk.read_date("class_0315_excel.xlsx")
    print("第一题的结果是：\n{}".format(res))
    wk.write_date("class_0315_excel.xlsx")

